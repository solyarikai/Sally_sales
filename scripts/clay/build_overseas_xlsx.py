#!/usr/bin/env python3
"""
Build 2-sheet XLSX from overseas workers data + decision makers.

Sheet 1 — Companies: Company, Domain, Non-US Employees, country1, country2, ...
Sheet 2 — People: Full Name, Job Title, Company, Domain (normalized), Location, LinkedIn

Domain normalization: strip http/https/www/trailing .com etc for VLOOKUP matching.

Usage:
  python3 build_overseas_xlsx.py \
    --companies exports/missouri_overseas_workers_v2.csv \
    --people exports/dm_people.json \
    --output ~/Downloads/Missouri_DM_Report.xlsx
"""

import argparse
import csv
import json
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill


def normalize_domain(domain):
    """Strip http/https/www/trailing TLD for matching."""
    if not domain:
        return ''
    d = domain.lower().strip()
    d = re.sub(r'^https?://', '', d)
    d = re.sub(r'^www\.', '', d)
    d = d.rstrip('/')
    # Remove trailing .com/.org/.net etc for core matching
    # But keep the full domain for display — only strip for the matching column
    return d


def core_domain(domain):
    """Extract core domain name (without TLD) for fuzzy matching."""
    d = normalize_domain(domain)
    # Remove TLD: .com, .org, .net, .io, .co, .us, etc.
    d = re.sub(r'\.[a-z]{2,6}$', '', d)
    # Remove country TLD prefix: .co.uk → .co
    d = re.sub(r'\.[a-z]{2,3}\.[a-z]{2}$', '', d)
    return d


# Location-to-country mapping (imported from JS module)
LOC_MAP = {
    'greater st. louis': 'United States', 'kansas city metropolitan area': 'United States',
    'missouri area': 'United States', 'greater chicago area': 'United States',
    'new york city metropolitan area': 'United States', 'atlanta metropolitan area': 'United States',
    'san francisco bay area': 'United States', 'dallas-fort worth metroplex': 'United States',
    'los angeles metropolitan area': 'United States', 'denver metropolitan area': 'United States',
    'greater phoenix area': 'United States', 'greater boston': 'United States',
    'washington dc-baltimore area': 'United States', 'detroit metropolitan area': 'United States',
    'italia': 'Italy', 'greater modena metropolitan area': 'Italy',
    'brasil': 'Brazil', 'são paulo e região': 'Brazil',
    'metro manila': 'Philippines', 'greater melbourne area': 'Australia',
    'greater sydney area': 'Australia', 'greater delhi area': 'India',
    'mumbai metropolitan region': 'India', 'greater cambridge area': 'United Kingdom',
    'greater paris metropolitan region': 'France', 'frankfurt/rhein-main': 'Germany',
    'schweiz': 'Switzerland', 'sverige': 'Sweden', 'madrid y alrededores': 'Spain',
    'türkiye': 'Turkey', 'nederland': 'Netherlands', 'czechia': 'Czech Republic',
    'méxico': 'Mexico',
    # Arabic
    'دبي الإمارات العربية المتحدة': 'United Arab Emirates',
    'الرياض السعودية': 'Saudi Arabia', 'القاهرة مصر': 'Egypt',
}

US_STATES = [
    'alabama','alaska','arizona','arkansas','california','colorado','connecticut',
    'delaware','florida','georgia','hawaii','idaho','illinois','indiana','iowa',
    'kansas','kentucky','louisiana','maine','maryland','massachusetts','michigan',
    'minnesota','mississippi','missouri','montana','nebraska','nevada',
    'new hampshire','new jersey','new mexico','new york','north carolina',
    'north dakota','ohio','oklahoma','oregon','pennsylvania','rhode island',
    'south carolina','south dakota','tennessee','texas','utah','vermont',
    'virginia','washington','west virginia','wisconsin','wyoming',
]

def to_country(location):
    if not location:
        return 'Unknown'
    norm = location.lower().strip()
    if norm in LOC_MAP:
        return LOC_MAP[norm]
    if norm == 'united states':
        return 'United States'
    for key, country in LOC_MAP.items():
        if key in norm:
            return country
    for state in US_STATES:
        if state in norm:
            return 'United States'
    return location.strip()


def main():
    parser = argparse.ArgumentParser(description='Build overseas DM report XLSX')
    parser.add_argument('--companies', required=True, help='Companies CSV (from overseas workers analysis)')
    parser.add_argument('--people', required=True, help='Decision makers JSON (from dm_people.json)')
    parser.add_argument('--output', default='overseas_dm_report.xlsx', help='Output XLSX path')
    args = parser.parse_args()

    # Read companies
    companies = []
    all_countries = set()
    with open(args.companies) as f:
        reader = csv.DictReader(f)
        for row in reader:
            breakdown = row.get('Non-US Breakdown', '')
            countries = {}
            if breakdown:
                for part in breakdown.split(','):
                    part = part.strip()
                    m = re.match(r'^(.+?):\s*(\d+)$', part)
                    if m:
                        countries[m.group(1).strip()] = int(m.group(2))
                        all_countries.add(m.group(1).strip())
            companies.append({
                'Company': row.get('Company', ''),
                'Domain': row.get('Domain', ''),
                'Domain_normalized': normalize_domain(row.get('Domain', '')),
                'Domain_core': core_domain(row.get('Domain', '')),
                'Total Employees': int(row.get('Total Employees', 0) or 0),
                'US Employees': int(row.get('US Employees', 0) or 0),
                'Non-US Employees': int(row.get('Non-US Employees', 0) or 0),
                'countries': countries,
            })

    # Sort countries by total count across all companies
    country_totals = {}
    for c in companies:
        for country, count in c['countries'].items():
            country_totals[country] = country_totals.get(country, 0) + count
    sorted_countries = sorted(country_totals.keys(), key=lambda x: -country_totals[x])

    # Build company lookup by normalized domain (for merging into People sheet)
    company_lookup = {}
    for c in companies:
        company_lookup[c['Domain_normalized']] = c

    # Read people
    with open(args.people) as f:
        people = json.load(f)

    print(f'Companies: {len(companies)}')
    print(f'People: {len(people)}')
    print(f'Countries: {len(sorted_countries)}')

    # Build XLSX
    wb = openpyxl.Workbook()

    # --- Sheet 1: Companies ---
    ws1 = wb.active
    ws1.title = 'Companies'

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # Count DMs per company domain
    dm_per_company = {}
    for p in people:
        d = normalize_domain(p.get('Company Domain', ''))
        dm_per_company[d] = dm_per_company.get(d, 0) + 1

    # Headers
    base_headers = ['Company', 'Domain', 'Domain (core)', 'DMs Found', 'Total Employees', 'US Employees', 'Non-US Employees']
    country_headers = sorted_countries  # country1, country2, etc.
    all_headers = base_headers + country_headers

    for col, header in enumerate(all_headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data
    for row_idx, co in enumerate(companies, 2):
        ws1.cell(row=row_idx, column=1, value=co['Company'])
        ws1.cell(row=row_idx, column=2, value=co['Domain'])
        ws1.cell(row=row_idx, column=3, value=co['Domain_core'])
        ws1.cell(row=row_idx, column=4, value=dm_per_company.get(co['Domain_normalized'], 0))
        ws1.cell(row=row_idx, column=5, value=co['Total Employees'])
        ws1.cell(row=row_idx, column=6, value=co['US Employees'])
        ws1.cell(row=row_idx, column=7, value=co['Non-US Employees'])
        for col_idx, country in enumerate(country_headers, 8):
            count = co['countries'].get(country, 0)
            if count > 0:
                ws1.cell(row=row_idx, column=col_idx, value=count)

    # Auto-width for first 6 columns
    for col in range(1, 7):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # --- Sheet 2: People (Decision Makers) ---
    ws2 = wb.create_sheet('Decision Makers')

    people_headers = ['# in Company', 'First Name', 'Last Name', 'Full Name', 'Email', 'Email Verified',
                      'Job Title', 'Company', 'Domain', 'Domain (core)', 'Location', 'Country', 'LinkedIn Profile',
                      'Overseas Employees', 'Overseas Breakdown',
                      'Top Country 1', 'Count 1', 'Top Country 2', 'Count 2', 'Top Country 3', 'Count 3']
    for col, header in enumerate(people_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Sequential numbering per company + merge overseas data
    seq_counters = {}
    for row_idx, p in enumerate(people, 2):
        location = p.get('Location', '')
        country = to_country(location.split(',')[-1].strip() if ',' in location else location)
        domain = p.get('Company Domain', '') or ''
        norm_domain = normalize_domain(domain)

        # Sequential number within company
        seq_counters[norm_domain] = seq_counters.get(norm_domain, 0) + 1

        # Get company overseas data via lookup
        co = company_lookup.get(norm_domain, {})
        co_countries = co.get('countries', {})
        top3 = sorted(co_countries.items(), key=lambda x: -x[1])[:3]

        # Split name if First/Last not provided
        first = p.get('First Name', '').strip()
        last = p.get('Last Name', '').strip()
        full = p.get('Full Name', '').strip()
        if not first and not last and full:
            name_parts = full.split()
            first = name_parts[0] if name_parts else ''
            last = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''

        ws2.cell(row=row_idx, column=1, value=seq_counters[norm_domain])
        ws2.cell(row=row_idx, column=2, value=first)
        ws2.cell(row=row_idx, column=3, value=last)
        ws2.cell(row=row_idx, column=4, value=full)
        ws2.cell(row=row_idx, column=5, value=p.get('Email', ''))
        ws2.cell(row=row_idx, column=6, value='Yes' if p.get('Email_Verified') else ('No' if p.get('Email') else ''))
        ws2.cell(row=row_idx, column=7, value=p.get('Job Title', ''))
        ws2.cell(row=row_idx, column=8, value=p.get('Company Name', p.get('Company Table Data', '')))
        ws2.cell(row=row_idx, column=9, value=domain)
        ws2.cell(row=row_idx, column=10, value=core_domain(domain))
        ws2.cell(row=row_idx, column=11, value=location)
        ws2.cell(row=row_idx, column=12, value=country)
        ws2.cell(row=row_idx, column=13, value=p.get('LinkedIn Profile', ''))
        ws2.cell(row=row_idx, column=14, value=sum(co_countries.values()))
        ws2.cell(row=row_idx, column=15, value=', '.join(f'{c}: {n}' for c, n in sorted(co_countries.items(), key=lambda x: -x[1])))
        for i in range(3):
            if i < len(top3):
                ws2.cell(row=row_idx, column=16 + i*2, value=top3[i][0])
                ws2.cell(row=row_idx, column=17 + i*2, value=top3[i][1])

    for col in range(1, 22):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['E'].width = 25

    # Save
    wb.save(args.output)
    print(f'\nXLSX saved to: {args.output}')
    print(f'  Sheet 1 "Companies": {len(companies)} rows, {len(all_headers)} columns')
    print(f'  Sheet 2 "Decision Makers": {len(people)} rows')
    print(f'  Domain matching column: "Domain (core)" — use for VLOOKUP between sheets')


if __name__ == '__main__':
    main()
